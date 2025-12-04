"""Populate the Epic-WAF Excel template with data from WAF YAML files."""

from __future__ import annotations

import argparse
import json
from copy import copy
from dataclasses import dataclass
from datetime import datetime
from pathlib import Path
from typing import Any, Dict, Iterable, List, Sequence

try:  # openpyxl drives all workbook manipulation.
    from openpyxl import load_workbook
    from openpyxl.styles import Alignment, PatternFill, Border, Side
    from openpyxl.worksheet.worksheet import Worksheet
    from openpyxl.pivot.cache import CalculatedItem
except ModuleNotFoundError as exc:  # pragma: no cover - import guard
    raise SystemExit("openpyxl is required. Install it with 'pip install openpyxl'.") from exc

try:  # PyYAML reads the mapping definition and WAF documents.
    import yaml
except ModuleNotFoundError as exc:  # pragma: no cover - import guard
    raise SystemExit("PyYAML is required. Install it with 'pip install pyyaml'.") from exc


DEFAULT_ROOT = Path(r"C:\temp\Epic_WAF_and_GoLive_Reviews-main")
DEFAULT_TOOLING_SUBDIR = "toolling"
DEFAULT_TEMPLATE_NAME = "Epic-WAF.xlsm"
DEFAULT_MAPPING_NAME = "yaml_mapping.yml"
DEFAULT_OUTPUT_SUBDIR = "excel"
DEFAULT_WAF_DIR = DEFAULT_ROOT / "WAF"
TARGET_SHEET = "Epic-RI-Checklist"
START_COLUMN_INDEX = 2  # Column B
START_ROW_INDEX = 3
HEADER_ROW_INDEX = START_ROW_INDEX - 1
COLUMN_M_INDEX = 13
COLUMN_A_FILL = PatternFill(start_color="FDFD1F", end_color="FDFD1F", fill_type="solid")
DATA_FILL = PatternFill(start_color="A6C9EC", end_color="A6C9EC", fill_type="solid")
BLANK_FILL = PatternFill(fill_type=None)
BORDER_THIN = Border(
    left=Side(border_style="thin", color="000000"),
    right=Side(border_style="thin", color="000000"),
    top=Side(border_style="thin", color="000000"),
    bottom=Side(border_style="thin", color="000000"),
)
BORDER_NONE = Border()

# Workaround for templates containing pivot cache calculated items with empty formulas
CalculatedItem.formula.allow_none = True


@dataclass
class ColumnSpec:
    key: str
    header: str
    order: int
    position: int  # preserves YAML iteration order to break ties consistently
    wrap_text: bool = False
    centered: bool = False


@dataclass
class CellPayload:
    text: str
    hyperlink: str | None = None


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument(
        "--root",
        type=Path,
        default=DEFAULT_ROOT,
        help="Base directory that contains the template, mapping file, and WAF folder",
    )
    parser.add_argument(
        "--template",
        type=Path,
        default=None,
        help="Path to the Epic-WAF.xlsm template (default: <root>/toolling/Epic-WAF.xlsm)",
    )
    parser.add_argument(
        "--mapping",
        type=Path,
        default=None,
        help="Path to yaml_mapping.yml (default: <root>/toolling/yaml_mapping.yml)",
    )
    parser.add_argument(
        "--waf-dir",
        type=Path,
        default=None,
        help="Directory containing WAF YAML files (default: <root>/WAF)",
    )
    parser.add_argument(
        "--output-dir",
        type=Path,
        default=None,
        help="Directory that will receive the generated workbook (default: <root>/excel)",
    )
    parser.add_argument(
        "--pattern",
        type=str,
        default="*.yml,*.yaml",
        help="Comma-separated glob patterns for YAML files (default: *.yml,*.yaml)",
    )
    return parser.parse_args()


def load_mapping(mapping_path: Path) -> List[ColumnSpec]:
    if not mapping_path.is_file():
        raise SystemExit(f"Mapping file not found: {mapping_path}")

    data = yaml.safe_load(mapping_path.read_text(encoding="utf-8")) or {}
    columns_section = data.get("epic_waf_excel_coloumns")
    if not isinstance(columns_section, dict):
        raise SystemExit("Mapping file must define 'epic_waf_excel_coloumns'.")

    specs: List[ColumnSpec] = []
    for position, (key, meta) in enumerate(columns_section.items()):
        if not isinstance(meta, dict):
            continue
        order = meta.get("order")
        if order is None:
            continue  # Only columns with an explicit order participate
        try:
            order_value = int(order)
        except (TypeError, ValueError) as exc:  # pragma: no cover - mapping issue
            raise SystemExit(f"Invalid order '{order}' for key '{key}'") from exc
        header = meta.get("name") or key
        wrap_flag = meta.get("wrap_text", 0)
        if isinstance(wrap_flag, str):
            wrap_text = wrap_flag.strip().lower() in {"1", "true", "yes"}
        else:
            wrap_text = bool(wrap_flag)

        centered_flag = meta.get("centered", 0)
        if isinstance(centered_flag, str):
            centered = centered_flag.strip().lower() in {"1", "true", "yes"}
        else:
            centered = bool(centered_flag)

        specs.append(
            ColumnSpec(
                key=key,
                header=str(header),
                order=order_value,
                position=position,
                wrap_text=wrap_text,
                centered=centered,
            )
        )

    if not specs:
        raise SystemExit("No ordered columns were defined in yaml_mapping.yml")

    specs.sort(key=lambda spec: (spec.order, spec.position))
    return specs


def iter_yaml_files(waf_dir: Path, patterns: Iterable[str]) -> List[Path]:
    files: List[Path] = []
    for pattern in patterns:
        if not pattern:
            continue
        files.extend(waf_dir.glob(pattern))
    return sorted({path.resolve() for path in files if path.is_file()})


def load_documents(files: Sequence[Path]) -> List[Dict[str, Any]]:
    documents: List[Dict[str, Any]] = []
    for file_path in files:
        raw = yaml.safe_load(file_path.read_text(encoding="utf-8"))
        if not isinstance(raw, dict):
            raise ValueError(f"File {file_path} does not contain a mapping at the root")
        documents.append(raw)
    return documents


def stringify_sequence(value: Any) -> str:
    if not isinstance(value, list):
        return serialize_scalar(value)
    entries: List[str] = []
    for item in value:
        if isinstance(item, str) and item.strip():
            entries.append(item.strip())
        elif item is not None:
            entries.append(str(item))
    return "\n".join(entries)


def format_link_field(value: Any) -> str:
    if isinstance(value, dict):
        link_text = value.get("link_name") or value.get("title") or ""
        link_target = value.get("link_uri") or value.get("link") or value.get("url") or ""
        if link_text and link_target:
            return f"{link_text} ({link_target})"
        if link_target:
            return str(link_target)
        if link_text:
            return str(link_text)
    return serialize_scalar(value)


def format_proactive_entries(value: Any) -> CellPayload:
    if not isinstance(value, list):
        return CellPayload(stringify_sequence(value))

    blocks: List[str] = []
    hyperlink: str | None = None
    for entry in value:
        if isinstance(entry, dict):
            title = (entry.get("title") or "").strip()
            description = (entry.get("description") or "").strip()
            link = (entry.get("link") or entry.get("link_uri") or "").strip()

            lines: List[str] = []
            if title:
                lines.append(title)
            if description and description != title:
                lines.append(description)

            line_text = "\n".join(lines) if lines else ""

            if link and title:
                if hyperlink is None:
                    hyperlink = link
                else:
                    link_display = f"{title} ({link})" if title else link
                    line_text = link_display if not lines else f"{line_text} ({link})"
            elif link:
                line_text = f"{line_text} ({link})" if line_text else link

            if line_text:
                blocks.append(line_text)
            continue

        if entry is not None:
            blocks.append(str(entry))

    text = "\n\n".join(blocks)
    return CellPayload(text=text, hyperlink=hyperlink)


def serialize_scalar(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, (str, int, float)):
        return str(value)
    if isinstance(value, bool):
        return "TRUE" if value else "FALSE"
    return json.dumps(value, ensure_ascii=False)


def column_value(document: Dict[str, Any], column: ColumnSpec) -> CellPayload:
    value = document.get(column.key)
    if value is None:
        return CellPayload("")
    if column.key in {"labels", "epic_resources", "epic_environment"}:
        return CellPayload(stringify_sequence(value))
    if column.key in {"arg_query", "manual_check"}:
        return CellPayload(format_link_field(value))
    if column.key == "proactive":
        return format_proactive_entries(value)
    return CellPayload(serialize_scalar(value))


def clear_existing_rows(sheet: Worksheet, column_count: int) -> int:
    last_row = find_last_populated_row(sheet, START_ROW_INDEX)
    if last_row < START_ROW_INDEX:
        return START_ROW_INDEX - 1
    for row in range(START_ROW_INDEX, last_row + 1):
        clear_data_row(sheet, row, column_count)
    return last_row


def find_last_populated_row(sheet: Worksheet, start_row: int) -> int:
    last_row = start_row - 1
    for row in range(sheet.max_row, start_row - 1, -1):
        cell = sheet.cell(row=row, column=START_COLUMN_INDEX)
        if cell.value not in (None, ""):
            last_row = row
            break
    return last_row


def copy_column_a_template(sheet: Worksheet, template_row: int, target_row: int) -> None:
    source = sheet.cell(row=template_row, column=1)
    target = sheet.cell(row=target_row, column=1)
    target.value = source.value
    if source.font is not None:
        target.font = copy(source.font)
    if source.fill is not None:
        target.fill = copy(source.fill)
    if source.border is not None:
        target.border = copy(source.border)
    if source.alignment is not None:
        target.alignment = copy(source.alignment)
    target.number_format = source.number_format
    if source.protection is not None:
        target.protection = copy(source.protection)
    target.fill = COLUMN_A_FILL
    target.border = BORDER_THIN


def replicate_column_a_validations(sheet: Worksheet, template_row: int, target_row: int) -> None:
    dv_list = getattr(sheet, "data_validations", None)
    if not dv_list:
        return
    source_coord = f"A{template_row}"
    target_coord = f"A{target_row}"
    for dv in dv_list.dataValidation:
        if any(source_coord in cell_range for cell_range in dv.ranges):
            dv.ranges.add(target_coord)


def style_data_row(sheet: Worksheet, row: int, column_count: int) -> None:
    for offset in range(column_count):
        cell = sheet.cell(row=row, column=START_COLUMN_INDEX + offset)
        cell.fill = DATA_FILL
        cell.border = BORDER_THIN


def clear_data_row(sheet: Worksheet, row: int, column_count: int) -> None:
    for offset in range(column_count):
        cell = sheet.cell(row=row, column=START_COLUMN_INDEX + offset)
        cell.value = None
        cell.hyperlink = None
        cell.fill = BLANK_FILL
        cell.border = BORDER_NONE


def apply_alignment(cell, wrap_text: bool, centered: bool) -> None:
    alignment = cell.alignment or Alignment()
    cell.alignment = alignment.copy(
        horizontal="center" if centered else "left",
        vertical="center" if centered else "top",
        wrap_text=bool(wrap_text),
    )


def populate_sheet(sheet: Worksheet, columns: Sequence[ColumnSpec], documents: Sequence[Dict[str, Any]]) -> None:
    column_count = len(columns)
    last_cleared_row = clear_existing_rows(sheet, column_count)

    for offset, spec in enumerate(columns):
        header_cell = sheet.cell(row=HEADER_ROW_INDEX, column=START_COLUMN_INDEX + offset)
        header_cell.value = spec.header
        apply_alignment(header_cell, spec.wrap_text, True)

    current_row = START_ROW_INDEX
    for document in documents:
        copy_column_a_template(sheet, START_ROW_INDEX, current_row)
        if current_row != START_ROW_INDEX:
            replicate_column_a_validations(sheet, START_ROW_INDEX, current_row)
        style_data_row(sheet, current_row, column_count)
        for offset, spec in enumerate(columns):
            payload = column_value(document, spec)
            cell = sheet.cell(row=current_row, column=START_COLUMN_INDEX + offset)
            cell.value = payload.text
            cell.hyperlink = payload.hyperlink
            if payload.hyperlink:
                cell.style = "Hyperlink"
            apply_alignment(cell, spec.wrap_text, spec.centered)
        current_row += 1

    # Clean up any leftover template rows from previous exports
    for row in range(current_row, last_cleared_row + 1):
        col_a = sheet.cell(row=row, column=1)
        col_a.value = None
        col_a.fill = BLANK_FILL
        col_a.border = BORDER_NONE
        clear_data_row(sheet, row, column_count)

    last_formatted_row = max(last_cleared_row, current_row - 1)
    if last_formatted_row >= START_ROW_INDEX:
        enforce_column_m_format(sheet, START_ROW_INDEX, last_formatted_row)


def enforce_column_m_format(sheet: Worksheet, start_row: int, end_row: int) -> None:
    for row in range(start_row, end_row + 1):
        cell = sheet.cell(row=row, column=COLUMN_M_INDEX)
        cell.border = BORDER_THIN
        cell.fill = DATA_FILL


def main() -> None:
    args = parse_args()

    root = args.root.expanduser().resolve()
    tooling_dir = (root / DEFAULT_TOOLING_SUBDIR).resolve()

    template_argument = args.template.expanduser() if args.template else None
    if template_argument is None:
        template_path = (tooling_dir / DEFAULT_TEMPLATE_NAME).resolve()
    elif template_argument.is_absolute():
        template_path = template_argument
    else:
        template_path = (root / template_argument).resolve()

    mapping_argument = args.mapping.expanduser() if args.mapping else None
    if mapping_argument is None:
        mapping_path = (tooling_dir / DEFAULT_MAPPING_NAME).resolve()
    elif mapping_argument.is_absolute():
        mapping_path = mapping_argument
    else:
        mapping_path = (root / mapping_argument).resolve()

    waf_dir = (args.waf_dir or DEFAULT_WAF_DIR).expanduser()
    if not waf_dir.is_absolute():
        waf_dir = (root / waf_dir).resolve()

    if args.output_dir is None:
        output_dir = (root / DEFAULT_OUTPUT_SUBDIR).resolve()
    else:
        output_dir_candidate = args.output_dir.expanduser()
        output_dir = output_dir_candidate if output_dir_candidate.is_absolute() else (root / output_dir_candidate).resolve()
    output_dir.mkdir(parents=True, exist_ok=True)

    if not template_path.is_file():
        raise SystemExit(f"Template workbook not found: {template_path}")
    if not waf_dir.is_dir():
        raise SystemExit(f"WAF directory not found: {waf_dir}")

    columns = load_mapping(mapping_path)
    patterns = [pattern.strip() for pattern in args.pattern.split(",")]
    files = iter_yaml_files(waf_dir, patterns)
    if not files:
        raise SystemExit(f"No YAML files found in {waf_dir}")

    documents = load_documents(files)
    if not documents:
        raise SystemExit("No YAML documents to export")

    workbook = load_workbook(template_path, keep_vba=True, data_only=False)
    if TARGET_SHEET not in workbook:
        raise SystemExit(f"Worksheet '{TARGET_SHEET}' not found in template")
    sheet = workbook[TARGET_SHEET]

    populate_sheet(sheet, columns, documents)

    timestamp = datetime.now().strftime("%Y-%m-%d_%H-%M-%S")
    output_path = output_dir / f"Epic Waf Checklist {timestamp}.xlsm"
    workbook.save(output_path)
    print(f"Workbook created: {output_path}")


if __name__ == "__main__":
    main()
